import matplotlib.pyplot as plt
import numpy as np
from pydub import AudioSegment
from openpyxl import Workbook

# ワークブックの新規作成と保存
wb = Workbook()
wb.save('myworkbook.xlsx')

# ワークブックの読み込み
from openpyxl import load_workbook

wb = load_workbook('myworkbook.xlsx')

ws = wb['Sheet'] 

# ボイスメモで収録したm4aファイルを読み込む
sounds = AudioSegment.from_wav("maou_se_system49.wav")

# 基本情報の表示
print(f'channel: {sounds.channels}')
print(f'frame rate: {sounds.frame_rate}')
print(f'duration: {sounds.duration_seconds} s')

# チャンネルが2 （ステレオ) の場合，交互にデータが入っているので，二つおきに読み出す。
# ただし，今回の場合はモノラルのはず。つまり，sounds.channels = 1
sig = np.array(sounds.get_array_of_samples())[::sounds.channels]
dt = 1.0/sounds.frame_rate # サンプリング時間

for i in range(len(sig)):
    ws.cell(row=i+1, column=1, value=sig[i])
wb.save('myworkbook.xlsx')     

# 時間アレイを作る
tms = 0.0 # サンプル開始時間を0にセット
tme = sounds.duration_seconds # サンプル終了時刻
tm = np.linspace(tms, tme, len(sig), endpoint=False) # 時間ndarrayを作成

# DFT
N = len(sig)
X = np.fft.fft(sig)
f = np.fft.fftfreq(N, dt) # Xのindexに対応する周波数のndarrayを取得

# データをプロット
fig, (ax01, ax02) = plt.subplots(nrows=2, figsize=(6, 8))
plt.subplots_adjust(wspace=0.0, hspace=0.6)

ax01.set_xlim(tms, tme)
ax01.set_xlabel('time (s)')
ax01.set_ylabel('x')
ax01.plot(tm, sig) # 入力信号

ax02.set_xlim(0, 8000)
ax02.set_xlabel('frequency (Hz)')

ax02.set_ylabel('|X|/N')
ax02.plot(f[0:N//2], np.abs(X[0:N//2])/N) # 振幅スペクトル

plt.show()